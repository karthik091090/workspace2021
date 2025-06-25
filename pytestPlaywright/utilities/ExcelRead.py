import openpyxl
import os

def getRowcount(path, sheetName):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetName]
    return sheet.max_row


def getColcount(path, sheetName):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetName]
    return sheet.max_column


def getCellData(path,sheetName,rowNo,colNo):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetName]
    return sheet.cell(row=rowNo,column=colNo).value


def setCellData(path, sheetName,rowNo,colNo,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetName]
    sheet.cell(row=rowNo,column=colNo).value=data
    workbook.save(path)
    print("saved")


base_dir=os.path.dirname(os.path.abspath(__file__))
print(base_dir)
path=os.path.join(base_dir,"..","Excel","logintest.xlsx")
print(path)
# path="./pytestPlaywright/Excel/logintest.xlsx"
sheetName="LoginSheet"
rows=getRowcount(path,sheetName)
cols=getColcount(path,sheetName)
print(rows,cols)

print(getCellData(path,sheetName,2,1))
setCellData(path,sheetName,1,5,"DOB")